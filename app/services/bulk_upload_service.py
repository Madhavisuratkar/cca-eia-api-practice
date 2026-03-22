import asyncio
from datetime import datetime
import io
from typing import Optional
import uuid

from openpyxl import load_workbook
from app.connections.cloud_s3_connect import check_file_exists_in_s3, fetch_s3_file, generate_upload_presigned_url, read_data_s3
from app.connections.custom_exceptions import CustomAPIException
from app.connections.env_config import CCA_UI, CHUNK_SIZE, LARGE_FILE_ROW_THRESHOLD
from app.connections.mongodb import get_collection
from app.models.policy_engine import PolicyEngine
from app.services.validation_service import read_udf_file_data_for_large
from app.utils.common_utils import BILLING_PARSERS, generate_user_name_from_email, extract_organization_from_email
from app.utils.constants import CollectionNames, MAX_EXCEL_ROW_LIMIT, RecommendationStatus, REQUIRED_HEADERS, AppName, LevelType
from app.connections.pylogger import log_message
from bson import ObjectId
import pandas as pd
from app.services.validation_service import regions_map
from sqlalchemy import func, or_
from app.services.portfolios_service import update_organization_instance_count


def policy_engine_sanity_check(db, user_email, provider, policy_engine):
    exists = (
        db.query(PolicyEngine.id)
        .filter(
            or_(
                PolicyEngine.user_email == user_email,
                PolicyEngine.user_email == ""
            ),
            func.lower(PolicyEngine.provider) == provider.lower(),
            func.lower(PolicyEngine.policy_name) == policy_engine.strip().lower(),
        )
        .first()
    )

    if not exists:
        raise CustomAPIException(
            message=(
                f"Policy engine '{policy_engine}' not found for provider '{provider}'"
            ),
            status_code=409,
        )

async def process_generate_upload_url(app_name: str, user_email: str, ipaddr: str, query, db):
    """Handles full logic for generating S3 upload URLs and managing portfolio records."""
    try:
        user_name = generate_user_name_from_email(user_email)

        # --- Validation ---
        if query.udf_file and app_name.upper() != AppName.EIA:
            raise CustomAPIException(
                status_code=400,
                message="udf_file can only be used when app_name is 'EIA'."
            )
        
        # ✅ Billing data validation
        if query.is_billing_data and app_name.upper() != AppName.CCA:
            raise CustomAPIException(
                status_code=400,
                message="is_billing_data can only be True when app_name is 'CCA'."
            )
        
        # ✅ Billing data validation
        if query.policy_engine and app_name.upper() != AppName.CCA:
            raise CustomAPIException(
                status_code=400,
                message="policy_engine should be app_name as 'CCA'."
            )

        if query.portfolio_id:
            await get_portfolio_by_id(query.portfolio_id, app_name)

        if query.policy_engine:
            policy_engine_sanity_check(db, user_email, query.provider, query.policy_engine)

        # --- Presigned URL Generation ---
        s3_key = udf_key = bulk_url = udf_url = ""

        if query.file_name:
            bulk_url, s3_key = generate_upload_presigned_url(
                app_name=app_name,
                user_name=user_name,
                file_name=query.file_name,
                sub_folder="input",
                file_type=query.file_type,
            )

        if query.udf_file:
            udf_url, udf_key = generate_upload_presigned_url(
                app_name=app_name,
                user_name=user_name,
                file_name=query.udf_file,
                sub_folder="udf",
                file_type=query.file_type,
            )

        # --- Portfolio update/create ---
        if query.udf_file:
            portfolio_id = await update_portfolio_file(query.portfolio_id, {"udf_key": udf_key})
        elif query.file_name and query.portfolio_id:
            portfolio_id = await update_portfolio_file(query.portfolio_id, {"s3_key": s3_key})

        else:
            portfolio_id = await handle_save_portfolio_with_s3(
                query.portfolioName,
                query.provider,
                query.headroom,
                s3_key,
                app_name,
                user_email,
                ipaddr,
                query.is_billing_data,
                query.policy_engine,
                query.created_for
            )

        # --- Build response ---
        response = {"portfolio_id": portfolio_id, "ErrorCode": 1}
        if bulk_url:
            response["presigned_url"] = bulk_url
        if udf_url:
            response["udf_presigned_url"] = udf_url

        return response
    except CustomAPIException:
        raise

    except Exception as e:
        log_message(LevelType.ERROR, f"{str(e)}, Details: process_generate_upload_url is failed", ErrorCode=-1)
        raise CustomAPIException(status_code=500, message="pre-signed url generation is failed", error_code=-1)


async def handle_save_portfolio_with_s3(portfolio_name, provider, headroom, s3_file_path, appname, user_email, ipaddr , is_billing_data = False, policy_engine=None, created_for=None):
    """"""
    try:
        now = datetime.utcnow()
        portfolio_collection = get_collection(CollectionNames.PORTFOLIOS)

        query = {
            "name": portfolio_name,
            "user_email": user_email,
            "cloud_provider": provider,
            "app_name": appname.upper(),
        }

        if is_billing_data:
            query["is_billing_data"] = True
        else:
            query["$or"] = [
                {"is_billing_data": False},
                {"is_billing_data": {"$exists": False}}
            ]

        existing = await portfolio_collection.find_one(query)
        if existing:
            portfolio_id = str(existing.get("_id"))
            log_message(LevelType.ERROR, f"Portfolio '{portfolio_name}'already exiting for app_name : {appname} : {provider}", ErrorCode=-1, portfolio_id=portfolio_id)
            raise CustomAPIException(status_code=400, message=f"Portfolio '{portfolio_name}'already exiting for appname : {appname} : {provider}")

        # Prevent creating portfolio with only UDF
        if not s3_file_path:
            raise CustomAPIException(
                status_code=400,
                message="Cannot create portfolio without excel file. Main file is required.",
                error_code=-1
            )
        record = {
                "user_email": user_email,
                "name": portfolio_name,
                "cloud_provider": provider,
                "headroom": headroom,
                "app_name": appname.upper(),
                "created_at": now,
                "uploaded_date": now,
                "s3_key": s3_file_path,
                "status": "Passed",
                "submittedForRecommendations": False,
                "recommendation_status": RecommendationStatus.TO_PROCESS,
                "is_cloud_cred": False,
                "is_billing_data" : is_billing_data,
                "ip": ipaddr,
                "policy_engine": policy_engine
            }
        if created_for:
            organization_data_collection = get_collection(CollectionNames.ORGANIZATION_DATA)
            org_doc = await organization_data_collection.find_one(
                {"organization": created_for, "app_name": appname.upper()}
            )

            if not org_doc:
                raise CustomAPIException(
                    status_code=400,
                    message=f"Organization '{created_for}' not found",
                    error_code=-1
                )
            record["created_for"] = created_for
        else:
            org = extract_organization_from_email(user_email)
            record["created_for"] = org
        result = await portfolio_collection.insert_one(record)
        portfolio_id = str(result.inserted_id)
        log_message(LevelType.INFO, "new portfolio created successfully", ErrorCode=1, portfolio_id=portfolio_id)
        return portfolio_id
    except CustomAPIException:
        raise

    except Exception as e:
        log_message(LevelType.ERROR, f"{str(e)}, Details: handle_save_portfolio_with_s3 is failed", ErrorCode=-1)
        raise CustomAPIException(status_code=500, message="portfolio creation failed", error_code=-1)


# --- Update main file (s3_key) for existing portfolio ---
async def update_portfolio_file(portfolio_id: str, update_data: dict):
    """
    Generic function to update a portfolio with given field(s).
    Example:
        await update_portfolio_file(portfolio_id, {"s3_key": s3_key})
        await update_portfolio_file(portfolio_id, {"udf_key": udf_key})
    """
    try:
        now = datetime.utcnow()
        portfolio_collection = get_collection(CollectionNames.PORTFOLIOS)

        # Add updated_at timestamp
        update_data["updated_at"] = now

        result = await portfolio_collection.update_one(
            {"_id": ObjectId(portfolio_id)},
            {"$set": update_data}
        )

        if result.matched_count == 0:
            raise CustomAPIException(
                status_code=404,
                message=f"No portfolio found with ID {portfolio_id}",
                error_code=-1
            )

        log_message(
            LevelType.INFO,
            f"Updated {', '.join(update_data.keys())} for portfolio_id: {portfolio_id}",
            portfolio_id=portfolio_id
        )

        return portfolio_id

    except CustomAPIException:
        raise
    except Exception as e:
        log_message(LevelType.ERROR, f"{str(e)}, Details: update_portfolio_file failed", ErrorCode=-1)
        raise CustomAPIException(status_code=500, message="Failed to update portfolio", error_code=-1)
    

async def get_portfolio_by_id(_id: str, app_name: str) -> dict:
    """
    Fetch portfolio by _id and app_name.
    Raises 404 CustomAPIException if not found.
    """
    try:
        portfolio_collection = get_collection(CollectionNames.PORTFOLIOS)
        portfolio_doc = await portfolio_collection.find_one({
            "_id": ObjectId(_id),
            "app_name": app_name.upper()
        })

        if not portfolio_doc:
            log_message(LevelType.ERROR, "Portfolio not found", ErrorCode=-1, portfolio_id=_id)
            raise CustomAPIException(
                status_code=404,
                message="Portfolio not found",
                error_code=-1
            )

        return portfolio_doc

    except CustomAPIException:
        raise
    except Exception as e:
        log_message(LevelType.ERROR, f"Error fetching portfolio: {str(e)}", ErrorCode=-1, portfolio_id=_id)
        raise CustomAPIException(
            status_code=500,
            message="Internal Server Error while fetching portfolio",
            error_code=-1
        )

async def insert_chunk(chunk_df, portfolio_id, chunk_counter, current_instance_collection, recommendation_tracking_collection, app_name, now):
    # Add required columns
    chunk_df["portfolio_id"] = portfolio_id
    chunk_df["created_at"] = now
    chunk_df["uploaded_date"] = now
    chunk_df["batch_id"] = chunk_counter

    records = chunk_df.to_dict(orient="records")

    # Insert current instances
    await current_instance_collection.insert_many(records)

    # Insert tracking document
    batch_doc = {
        "batch_id": chunk_counter,
        "portfolio_id": portfolio_id,
        "app_name": app_name,
        "record_count": len(records),
        "created_at": now,
        "recommendation_status": RecommendationStatus.QUEUE
    }
    await recommendation_tracking_collection.insert_one(batch_doc)
    return len(records)


def validate_required_headers(app_name: str, df, portfolio_name: str, s3_key: str, portfolio_id: str = None):
    required_headers = REQUIRED_HEADERS[app_name]

    # Normalize DataFrame columns (case-insensitive)
    df.columns = [c.strip().lower() for c in df.columns]
    required_headers_lower = [h.lower() for h in required_headers]

    missing_headers = [h for h in required_headers_lower if h not in df.columns]
    if missing_headers:
        message = (
            f"Missing required headers in Excel for portfolio '{portfolio_name}': "
            f"[{', '.join(missing_headers)}]. Please download the template to get exact headers."
        )
        log_message(LevelType.ERROR, f"{message} for s3_key : {s3_key}", ErrorCode=-1, portfolio_id=portfolio_id)
        raise CustomAPIException(status_code=400, message=message)
        

# --- API callable service ---
async def start_cost_advice_recommendation(portfolio_id: str, app_name: str):
    try:
        if not ObjectId.is_valid(portfolio_id):  # format: 24-char hex or 12-byte input
            raise CustomAPIException(
                status_code=400,
                message=f"Invalid portfolio_id '{portfolio_id}' Must be a valid 24-character hex string."
            ) 
        portfolio_collection = get_collection(CollectionNames.PORTFOLIOS)
        projection = {
            "_id": 1,
            "name": 1,
            "recommendation_status": 1,
            "user_email": 1,
            "s3_key": 1,
            "cloud_provider" : 1,
            "is_billing_data" : 1
        }

        # Add udf_key if app_name is EIA
        if app_name.upper() == "EIA":
            projection["udf_key"] = 1

        portfolio = await portfolio_collection.find_one(
            {"_id": ObjectId(portfolio_id), "app_name": app_name},
            projection
        )
        if not portfolio:
            log_message(LevelType.ERROR, "portfolio not found", ErrorCode=-1, portfolio_id=portfolio_id)
            raise CustomAPIException(status_code=404, message=f"Portfolio '{portfolio_id}' not found for app '{app_name}'")

        portfolio_name = portfolio.get("name")
        s3_key = portfolio.get("s3_key")
        udf_key = portfolio.get("udf_key")
        is_billing_data = portfolio.get("is_billing_data", False)
        provider = portfolio.get("cloud_provider")
        recommendation_status = portfolio.get("recommendation_status")
        if  recommendation_status in [RecommendationStatus.QUEUE,RecommendationStatus.IN_PROGRESS]:
            log_message(LevelType.ERROR, f"recommendation_status is {recommendation_status}", ErrorCode=-1, portfolio_id=portfolio_id)
            raise CustomAPIException(status_code=400, message=f"Recommendation already in progress for portfolio '{portfolio_name}'")
        
        # --- Validate S3 file existence ---
        await validate_s3_file_exists(s3_key, portfolio_id, portfolio_name)
        log_message(LevelType.INFO, f"s3 file : {s3_key} is exists and is_billing_data : {is_billing_data} - cloud_provider : {provider}", ErrorCode=1, portfolio_id=portfolio_id)

        s3_obj = read_data_s3(s3_key)
        file_stream = s3_obj["Body"].read()
        df = pd.read_excel(io.BytesIO(file_stream))

        # 🔹 Drop unnamed or empty columns
        df = df.loc[:, ~df.columns.str.match(r"Unnamed", case=False)]

        if df.empty:
            raise CustomAPIException(status_code=400, message=f"No data found in Excel for portfolio '{portfolio_name}'")

        # 🔹 Replace any NA/NaN values with empty string
        df = df.fillna("")
        # 🔹 Keep a copy of df before validation
        df_copy = df.copy()
        
        if not is_billing_data:
            validate_required_headers(app_name.upper(), df, portfolio_name, s3_key, portfolio_id)
            # 🔹 Restore 'uuid' from the copy if missing
            if 'uuid' not in df.columns and 'uuid' in df_copy.columns:
                df['uuid'] = df_copy['uuid'].fillna("")  # keep empty strings where it was missing
            elif 'uuid' in df.columns:
                # Fill any missing UUIDs in existing column with empty string
                df['uuid'] = df['uuid'].fillna("")
        else:
            df.columns = [c.strip().lower() for c in df.columns]
            parser_func = BILLING_PARSERS[provider]
            df = parser_func(df, provider, portfolio_id, regions_map)
            if df.empty:
                log_message(LevelType.ERROR, "No instance SKUs found from billing", ErrorCode=-1,  portfolio_id=portfolio_id)
                raise CustomAPIException(status_code=400, message=f"No instance SKUs found from billing for portfolio '{portfolio_name}'")


        if udf_key:
            udf_exists = await check_file_exists_in_s3(udf_key)
            if not udf_exists:
                log_message(LevelType.ERROR, f"udf file : {udf_key} is not exists", ErrorCode=-1, portfolio_id=portfolio_id)
                raise CustomAPIException(
                    status_code=404,
                    message=f"udf data file not found for portfolio '{portfolio_name}' at {s3_key}"
                )
            log_message(LevelType.INFO, f"udf file : {udf_key} is exists", ErrorCode=1, portfolio_id=portfolio_id)

            udf_file = fetch_s3_file(udf_key)
            udf_data, udf_message = await read_udf_file_data_for_large(udf_file)
            if not udf_data:
                message = f"Own metrics data: {udf_message}"
                log_message(LevelType.INFO, f"{message} for file : {udf_key}",ErrorCode=1)

        
        current_instance_collection = get_collection(CollectionNames.CURRENT_INSTANCES)
        recommendation_tracking_collection = get_collection(CollectionNames.RECOMMENDATION_TRACKING)            

        log_message(LevelType.INFO, f"Loaded {len(df)} rows from {s3_key}", ErrorCode=1, portfolio_id=portfolio_id)

        # --- Drop existing portfolio data in current_instance_collection ---
        delete_result = await current_instance_collection.delete_many({"portfolio_id": portfolio_id})
        log_message(LevelType.INFO, f"Deleted {delete_result.deleted_count} old records for portfolio_id {portfolio_id}", ErrorCode=1, portfolio_id=portfolio_id)

        # --- Drop existing portfolio tracking data in recommendation_tracking_collection ---
        delete_tracking_result = await recommendation_tracking_collection.delete_many({"portfolio_id": portfolio_id})
        log_message(LevelType.INFO, f"Deleted {delete_tracking_result.deleted_count} old records for portfolio_id {portfolio_id}", ErrorCode=1, portfolio_id=portfolio_id)

        # ---- Process chunks ----
        now = datetime.utcnow()
        total_rows = len(df)
        chunk_counter = 0
        total_inserted = 0

        # Create async tasks for each chunk
        tasks = []
        for start_idx in range(0, total_rows, CHUNK_SIZE):
            chunk_counter += 1
            chunk_df = df.iloc[start_idx:start_idx + CHUNK_SIZE].copy()
            tasks.append(
                insert_chunk(chunk_df, portfolio_id, chunk_counter,
                            current_instance_collection, recommendation_tracking_collection,
                            app_name, now)
            )

        # Run tasks concurrently
        results = await asyncio.gather(*tasks)

        total_inserted = sum(results)
        log_message(LevelType.INFO, f"Completed inserting {total_inserted} records for portfolio '{portfolio_name}'", ErrorCode=1, portfolio_id=portfolio_id)

        # Update portfolio recommendation_status as QUEUE
        update_fields = {
            "recommendation_status": RecommendationStatus.QUEUE,
            "is_large_data": True,
            "updated_at": datetime.utcnow(),
            "current_instances_count": total_inserted
        }

        # Include udf_data if available
        if udf_key and udf_data:
            update_fields["udf"] = udf_data

        # Update portfolio document
        await portfolio_collection.update_one(
            {"_id": ObjectId(portfolio_id)},
            {"$set": update_fields}
        )

        inc_result = await update_organization_instance_count(
            portfolio_id=portfolio_id,
            new_current_instance_count=total_inserted,
            app_name=app_name.upper()
        )

        # Log it
        log_message(LevelType.INFO, f"[ORG_UPDATE] {inc_result['message']}", ErrorCode=1)

        response_send = " via email"
        if "ccr" in CCA_UI:
            response_send = ""

        return {
            "status": "success",
            "message": (
                f"Cost advice recommendation initiated for portfolio '{portfolio_name}'. "
                f"You will be notified{response_send} once the recommendation is completed."
            ),
            "portfolio_id": portfolio_id,
            "s3_key": s3_key,
            "ErrorCode": 1
        }
    except CustomAPIException:
        raise
    except Exception as e:
        msg = f"Unexpected error in start_cost_advice_recommendation: {str(e)}"
        log_message(LevelType.ERROR, msg, ErrorCode=-1, portfolio_id=portfolio_id)
        raise CustomAPIException(
            status_code=500,
            message="Internal Server Error while processing cost advice recommendation",
            error_code=-1
        )


async def validate_s3_file_exists(s3_path: Optional[str], portfolio_id: str, portfolio_name: str):
    """
    Check if the given S3 path exists. Raises CustomAPIException if not.
    """
    if not s3_path:
        log_message(LevelType.ERROR, f"s3 file : {s3_path} is not available from portfolio", ErrorCode=-1, portfolio_id=portfolio_id)
        raise CustomAPIException(status_code=400, message=f"No source data (s3_path) found for portfolio '{portfolio_name}'")
    
    file_exists = await check_file_exists_in_s3(s3_path)
    if not file_exists:
        log_message(LevelType.ERROR, f"s3 file : {s3_path} does not exist", ErrorCode=-1, portfolio_id=portfolio_id)
        raise CustomAPIException(status_code=404, message=f"S3 data file not found for portfolio '{portfolio_name}' at {s3_path}")


# hanldle row count
async def get_excel_row_count(portfolio_id: str, app_name: str):
    """
    Get row count of an Excel file from S3 using openpyxl in read-only mode.
    """
    s3_path = None  # ensure defined for except scope [web:2][web:14]
    wb = None
    try:
         # 1) Validate ObjectId early
        if not ObjectId.is_valid(portfolio_id):  # format: 24-char hex or 12-byte input
            raise CustomAPIException(
                status_code=400,
                message=f"Invalid portfolio_id '{portfolio_id}'."
            ) 

        portfolio_collection = get_collection(CollectionNames.PORTFOLIOS)
        projection = {
            "_id": 1,
            "name": 1,
            "recommendation_status": 1,
            "user_email": 1,
            "s3_key": 1,
            "cloud_provider" : 1
        }
    
        portfolio = await portfolio_collection.find_one(
            {"_id": ObjectId(portfolio_id), "app_name": app_name},
            projection
        )
        if not portfolio:
            log_message(LevelType.ERROR, "Portfolio not found", ErrorCode=-1, portfolio_id=portfolio_id)
            raise CustomAPIException(status_code=404, message=f"Portfolio '{portfolio_id}' not found for app '{app_name}'")
    
        portfolio_name = portfolio.get("name")
        s3_path = portfolio.get("s3_key")

        # --- Validate S3 file existence ---
        await validate_s3_file_exists(s3_path, portfolio_id, portfolio_name)


        log_message(LevelType.INFO, f"Downloading file from {s3_path}", ErrorCode=1)

        # Stream download from S3
        s3_obj = read_data_s3(s3_path)
        file_stream = s3_obj["Body"].read()

        # Fast row count with openpyxl
        wb = load_workbook(filename=io.BytesIO(file_stream), read_only=True)
        ws = wb.worksheets[0]  # Use only first sheet
        row_count = ws.max_row - 1

        if row_count > MAX_EXCEL_ROW_LIMIT:
            msg = f"Excel sheet '{ws.title}' exceeds maximum allowed rows ({MAX_EXCEL_ROW_LIMIT}). Found {row_count} rows."
            log_message(LevelType.ERROR, msg, ErrorCode=-1, portfolio_id=portfolio_id)
            raise CustomAPIException(status_code=400, message=msg)

        # Use threshold constant
        is_large_file = row_count >= LARGE_FILE_ROW_THRESHOLD 

        log_message(LevelType.INFO, f"Row count for {s3_path}: {row_count}", ErrorCode=1)

        return {
            "is_large_file": is_large_file, "ErrorCode": 1
        }
    
    except CustomAPIException:
        raise

    except Exception as e:
        log_message(LevelType.ERROR, f"Failed to read Excel file {s3_path}: {str(e)}", ErrorCode=-1)
        raise CustomAPIException(
            status_code=500,
            message=f"Failed to read Excel file {s3_path}: {str(e)}"
        )


async def get_recommendation_progress_service(portfolio_id, app_name):
    """
    Service function to fetch recommendation progress details for a portfolio.
    """
    try:
        portfolio = await get_portfolio_by_id(portfolio_id, app_name)
        # Prepare default progress dict
        progress = {
            "portfolio_id": portfolio_id,
            "completion_percentage": 0,
            "completed_batches": 0,
            "failed_batches": 0,
            "queued_batches": 0,
            "total_batches": 0,
            "ErrorCode": 1
        }

        status = portfolio.get("recommendation_status")
        if status in [RecommendationStatus.QUEUE]:
            log_message(LevelType.INFO, "Portfolio recommendation progress fetched successfully", ErrorCode=1, portfolio_id=portfolio_id)
            return progress

        if status == RecommendationStatus.COMPLETED or status == RecommendationStatus.FAILED:
            progress["completion_percentage"] = 100
            log_message(LevelType.INFO, f"Portfolio recommendation progress fetched successfully for the status : {status}", ErrorCode=1, portfolio_id=portfolio_id)
            return progress

        collection = get_collection(CollectionNames.RECOMMENDATION_TRACKING)
        cursor = collection.find({"portfolio_id": portfolio_id})
        batches = await cursor.to_list(length=None)

        if not batches:
            log_message(LevelType.ERROR, "No input data found to fetch recommendation progress", ErrorCode=-1)
            return progress

        progress["total_batches"] = max(batch["batch_id"] for batch in batches)
        progress["completed_batches"] = sum(1 for batch in batches if batch["recommendation_status"].upper() == RecommendationStatus.COMPLETED)
        progress["failed_batches"] = sum(1 for batch in batches if batch["recommendation_status"].upper() == RecommendationStatus.FAILED)
        progress["queued_batches"] = sum(1 for batch in batches if batch["recommendation_status"].upper() == RecommendationStatus.QUEUE)

        if progress["total_batches"] > 0:
            completion = ((progress["completed_batches"] + progress["failed_batches"]) / progress["total_batches"]) * 100

            # Deduct 1 batch worth percentage
            if progress["total_batches"] > 0:
                completion -= (100 / progress["total_batches"])

            progress["completion_percentage"] = round(max(min(completion, 99.99), 0), 2)

        log_message(LevelType.INFO, "Portfolio recommendation progress fetched successfully", ErrorCode=1, portfolio_id=portfolio_id)
        return progress

    except CustomAPIException:
        raise
    except Exception as e:
        log_message(LevelType.ERROR, f"Facing issue in fetching progress: {str(e)}", ErrorCode=-1)
        raise CustomAPIException(status_code=500, message=f"Failed to fetch progress: {str(e)}")
