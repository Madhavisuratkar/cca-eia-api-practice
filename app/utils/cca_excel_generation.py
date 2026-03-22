from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from app.connections.pylogger import log_message
from app.utils.constants import LevelType

WHITE_BORDER = Border(
    left=Side(style="thin", color="FFFFFF"),
    right=Side(style="thin", color="FFFFFF"),
    top=Side(style="thin", color="FFFFFF"),
    bottom=Side(style="thin", color="FFFFFF"),
)


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def style_cell(cell, fill=None, font=None, align=None, border=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border


def safe_round(value, decimals=2):
    try:
        if value is None or str(value).strip() == "":
            return value
        val = float(value)
        return round(val, decimals)
    except (ValueError, TypeError):
        return value


def format_savings(value, decimals=2):
    try:
        if value is None or str(value).strip() == "":
            return value
        val = float(value)
        if val < 0:
            return "EIA Recommended"
        return round(val, decimals)
    except (ValueError, TypeError):
        return value


def set_col_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if cell and cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[col_letter].width = max_len + 2
    ws.column_dimensions["A"].width = 15


# ============================================================
# 1) CREATE MAIN SHEET HEADER
# ============================================================

def create_main_header(ws, total_columns, header_fill):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    cell = ws["A1"]
    cell.value = "EPYC Cloud Cost Advisory Recommendations"

    style_cell(cell,
               fill=header_fill,
               font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
               align=Alignment(horizontal="center", vertical="center"))

    for col in range(1, total_columns + 1):
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).border = WHITE_BORDER

    # Increase row 1 height
    ws.row_dimensions[1].height = 25


# ============================================================
# 2) CREATE TABLE HEADERS (Rows 2 & 3)
# ============================================================

def create_table_headers(ws, header_fill):

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(wrap_text=True, horizontal="center", vertical="center")

    main_cols = [
        "Region", "Current Instance", "Current Monthly Cost ($)", "Current Annual Cost ($)",
        "UUID/Instance Name", "Cloud", "Quantity", "Pricing Model",
        "Current vCPU(s)", "Remark"
    ]

    for idx, text in enumerate(main_cols, start=1):
        ws.merge_cells(start_row=2, start_column=idx, end_row=3, end_column=idx)
        cell = ws.cell(2, idx, value=text)
        style_cell(cell, header_fill, header_font, header_align, border=WHITE_BORDER)
        # Apply border to row 3 as well for merged cells
        ws.cell(3, idx).fill = header_fill
        ws.cell(3, idx).border = WHITE_BORDER

    rec_groups = [
        (11, "Hourly Cost Optimization"),
        (18, "Modernize"),
        (25, "Modernize & Downsize")
    ]

    for col, title in rec_groups:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 6)
        cell = ws.cell(2, col, value=title)
        style_cell(cell, header_fill, header_font, header_align, border=WHITE_BORDER)
        # Apply border to all cells in the merged range
        for c in range(col, col + 7):
            ws.cell(2, c).fill = header_fill
            ws.cell(2, c).border = WHITE_BORDER

    detailed_headers = [
        "Instance", "vCPU(s)", "Monthly Cost ($)", "Annual Cost ($)",
        "Annual Savings ($)", "Savings (%)", "Performance Improvement"
    ] * 3

    for col, text in enumerate(detailed_headers, start=11):
        cell = ws.cell(3, col, value=text)
        style_cell(cell, header_fill, header_font, header_align, border=WHITE_BORDER)
    
    # Adjust row heights for headers
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20


# ============================================================
# 3) SORT DATA BY RECOMMENDATIONS (GREEN FIRST, WHITE LAST)
# ============================================================

def _has_green_recommendation(item):
    """Returns True if any recommendation has positive annual savings (green cell)."""
    recs = item.get("data", {}).get("recommendations", [])
    for rec in recs:
        try:
            if float(rec.get("annualSavings", 0)) > 0:
                return True
        except (ValueError, TypeError):
            pass
    return False


def sort_data_green_first(data):
    """Sort data so rows with green recommendations come first, white (no recommendations) last."""
    return sorted(data, key=lambda x: (0 if _has_green_recommendation(x) else 1))


# ============================================================
# 4) WRITE DATA ROWS
# ============================================================

def write_data_rows(ws, data, green_fill):
    for item in data:
        current = item["data"]["currentPlatform"]
        recs = item["data"]["recommendations"]

        row = [
            current["zone"], current["instanceType"], safe_round(current["monthlyCost"]),
            safe_round(current["annualCost"]), item["id"], current["cspProvider"],
            current["numberOfInstances"], current["pricingModel"],
            safe_round(current["vCPU"]), current["status"]
        ]

        for rec in recs:
            row.extend([
                rec["instanceType"], safe_round(rec["vCPU"]), safe_round(rec["monthlyCost"]),
                safe_round(rec["totalCost"]), format_savings(rec["annualSavings"]),
                safe_round(rec["savingsInPercentage"]), safe_round(rec["perf"])
            ])

        ws.append(row)

        last_row = ws.max_row
        for col in range(1, len(row) + 1):
            ws.cell(row=last_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            
            # Apply thousand separator formatting to Cost, Annual Cost, Annual Savings cols
            if col in [3, 4] or (col >= 11 and (col - 11) % 7 in [2, 3, 4]):
                ws.cell(row=last_row, column=col).number_format = '#,##0.00'

        for idx, rec in enumerate(recs):
            try:
                if float(rec["annualSavings"]) > 0:
                    instance_col = 11 + (idx * 7)
                    ws.cell(last_row, instance_col).fill = green_fill
            except:
                pass


# ============================================================
# 4) ADD GRAND TOTAL ROW
# ============================================================

def write_grand_total(ws, grand_total):
    row = [
        "Grand Total", "", safe_round(grand_total["Current Monthly Cost"]),
        safe_round(grand_total["Annual Cost"]), "", "", grand_total["Number of Instances"],
        "", "", "", "", "", safe_round(grand_total["Monthly Cost I"]),
        safe_round(grand_total["Annual Cost I (perf scaled)"]),
        safe_round(grand_total["Annual Savings I"]), safe_round(grand_total["hSavingsInPercentage"]),
        safe_round(grand_total["Perf Enhancement I"]), "", "",
        safe_round(grand_total["Monthly Cost II"]), safe_round(grand_total["Annual Cost II (perf scaled)"]),
        safe_round(grand_total["Annual Savings II"]), safe_round(grand_total["mSavingsInPercentage"]),
        safe_round(grand_total["Perf Enhancement II"]), "", "",
        safe_round(grand_total["Monthly Cost III"]),
        safe_round(grand_total["Annual Cost III (perf scaled)"]),
        safe_round(grand_total["Annual Savings III"]), safe_round(grand_total["mdSavingsInPercentage"]),
        safe_round(grand_total["Perf Enhancement III"]),
    ]

    ws.append(row)
    last_row = ws.max_row

    for col in range(1, len(row) + 1):
        cell = ws.cell(row=last_row, column=col)
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = Border()
        
        # Apply thousand separator formatting to numeric total columns
        if col in [3, 4] or (col >= 13 and (col - 13) % 7 in [0, 1, 2]):
            cell.number_format = '#,##0.00'


# ============================================================
# 5) ADD NOTE ROW
# ============================================================

def add_note(ws):
    note_row = ws.max_row + 2
    ws.merge_cells(f"A{note_row}:F{note_row}")

    cell = ws[f"A{note_row}"]
    cell.value = "Note: Green color instances indicate positive savings."

    style_cell(cell,
               font=Font(name="Calibri", size=11, color="0000FF"),
               align=Alignment(horizontal="left", vertical="center"))
    
    # Add second note
    note_row2 = note_row + 1
    ws.merge_cells(f"A{note_row2}:F{note_row2}")
    
    cell2 = ws[f"A{note_row2}"]
    cell2.value = "Sizing instances - matching resources to actual demand (if Applicable)"
    
    style_cell(cell2,
               font=Font(name="Calibri", size=11, color="0000FF"),
               align=Alignment(horizontal="left", vertical="center"))


# ============================================================
# 6) SUMMARY SHEET
# ============================================================

def create_summary_sheet(wb, grand_total, header_fill):
    ws = wb.create_sheet("Total Annual Savings")

    white = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:A2")
    style_cell(ws["A1"], header_fill, white, center, border=WHITE_BORDER)
    ws["A1"].value = "Current Cost"

    groups = {"B": "Hourly Cost Optimization", "D": "Modernize", "F": "Modernize & Downsize"}

    for col, title in groups.items():
        ws.merge_cells(f"{col}1:{chr(ord(col)+1)}1")
        cell = ws[f"{col}1"]
        cell.value = title
        style_cell(cell, header_fill, white, center, border=WHITE_BORDER)

    sub = ["Total Cost", "Total Savings"] * 3
    for idx, col in enumerate(["B","C","D","E","F","G"]):
        style_cell(ws[f"{col}2"], header_fill, white, center, border=WHITE_BORDER)
        ws[f"{col}2"].value = sub[idx]

    data = [
        safe_round(grand_total["Annual Cost"]),
        safe_round(grand_total["Annual Cost I (perf scaled)"]), safe_round(grand_total["Annual Savings I"]),
        safe_round(grand_total["Annual Cost II (perf scaled)"]), safe_round(grand_total["Annual Savings II"]),
        safe_round(grand_total["Annual Cost III (perf scaled)"]), safe_round(grand_total["Annual Savings III"]),
    ]

    for col, value in zip(["A","B","C","D","E","F","G"], data):
        cell = ws[f"{col}3"]
        cell.value = float(value)
        cell.number_format = '"$ "#,##0.00'

    set_col_widths(ws)


# ============================================================
# 7) LEGAL DISCLAIMER SHEET
# ============================================================

def create_legal_sheet(wb):
    ws = wb.create_sheet("Legal Disclaimer")

    ws.merge_cells("A1:Q1")
    cell = ws["A1"]
    cell.value = (
        "Disclaimer: THE MATERIALS PROVIDED THROUGH THIS TOOL ARE PROVIDED 'AS IF', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE."
    )
    cell.font = Font(name="Calibri", size=11)
    cell.alignment = Alignment(wrap_text=False)

    ws.row_dimensions[1].height = 25

    ws.merge_cells("A3:G3")
    ws["A3"] = "Cloud instance recommendation generated using AMD EPYC Cloud Cost Advisor"

    ws.merge_cells("A5:G5")
    current_year = datetime.now().year
    ws["A5"] = f"Copyright - {current_year} Advanced Microdevices Inc."

    ws.merge_cells("A7:H7")
    ws["A7"] = "For terms: https://www.amd.com/en/legal/copyright.html"

    for col in "ABCDEFGHIJKLMNOPQ":
        ws.column_dimensions[col].width = 20


# ============================================================
# MAIN FUNCTION (JSON INPUT DIRECTLY)
# ============================================================

def generate_excel_from_json(input_json, output_path):
    try:
        # log_message(LevelType.INFO, f"Generating Excel from JSON: {input_json}")
        log_message(LevelType.INFO, f"Output path: {output_path}")
        
        data = input_json["data"]
        grand_total = input_json["grandTotal"]

        wb = Workbook()
        ws = wb.active
        ws.title = "Recommended-Instance"

        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        create_main_header(ws, 31, header_fill)
        create_table_headers(ws, header_fill)

        ws.freeze_panes = "E4"

        sorted_data = sort_data_green_first(data)
        write_data_rows(ws, sorted_data, green_fill)
        write_grand_total(ws, grand_total)
        add_note(ws)
        set_col_widths(ws)

        create_summary_sheet(wb, grand_total, header_fill)
        create_legal_sheet(wb)

        wb.save(output_path)
        log_message(LevelType.INFO, "Excel successfully generated!")
    except Exception as e:
        log_message(LevelType.ERROR, f"Error generating Excel: {e}", ErrorCode=-1)
