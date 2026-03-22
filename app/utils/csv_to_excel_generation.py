import csv
from app.connections.pylogger import log_message
from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.cell.cell import MergedCell
from app.utils.constants import LevelType, NEGATIVE_REPLACEMENT_TEXT

def create_summary_sheet(wb, sheet, run_type):
    """
    Creates a new summary sheet using values from the 'Grand Total' row.
    """
    try:
        if run_type == "EIA":
            annual_sav = {
                "Current Cost": "Current Monthly Price",
                "OPTIMAL Total Cost": "Monthly Price I",
                "OPTIMAL Total Savings": "Monthly Savings I",
                "GOOD Total Cost": "Monthly Price II",
                "GOOD Total Savings": "Monthly Savings II"
            }
        elif run_type == "CCA":
            annual_sav ={
                "Current Cost": "Annual Cost",
                "OPTIMAL Total Cost": "Annual Cost I (perf scaled)",
                "OPTIMAL Total Savings": "Annual Savings I",
                "BEST Total Cost": "Annual Cost II (perf scaled)",
                "BEST Total Savings": "Annual Savings II",
                "GOOD Total Cost": "Annual Cost III (perf scaled)",
                "GOOD Total Savings": "Annual Savings III"
            }
        else:
            return None

        # Find relevant columns
        header = [cell.value for cell in sheet[1]]
        target_columns = {col: idx + 1 for idx, col in enumerate(header)}

        # Find the row with 'Grand Total' and extract values
        grand_total_values = {}
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value and "Grand Total" in str(row[0].value):
                for key, col_name in annual_sav.items():
                    col_idx = target_columns.get(col_name)
                    if col_idx:
                        row_val = row[col_idx - 1].value
                        if run_type == "EIA":
                            row_val = row_val * 12
                        grand_total_values[key] = row_val
                        
                        # grand_total_values[key] = f"${round(row_val,2)}"
                break  # Stop after finding the first "Grand Total" row

        # Set headers in summary sheet
        if run_type == "EIA":
            headers_first_row = ["Current Cost", "OPTIMAL", "", "GOOD", ""]
            headers_second_row = ["Current Cost", "Total Cost", "Total Savings", "Total Cost", "Total Savings"]
        if run_type == "CCA":
            headers_first_row = ["Current Cost", "Hourly Cost Optimization", "", "Modernize", "", "Modernize & Downsize", ""]
            headers_second_row = ["Current Cost", "Total Cost", "Total Savings", "Total Cost", "Total Savings", "Total Cost", "Total Savings"]

        if grand_total_values:
            summary_sheet = wb.create_sheet(title="Total Annual Savings")
            summary_sheet.append(headers_first_row)
            summary_sheet.append(headers_second_row)
            summary_sheet.append([grand_total_values.get(key, "") for key in annual_sav.keys()])

        # Merge header cells for grouping
        summary_sheet.merge_cells("A1:A2")
        summary_sheet.merge_cells("B1:C1")
        if run_type == "EIA":
            summary_sheet.merge_cells("D1:E1")  # GOOD section
        elif run_type == "CCA":
            summary_sheet.merge_cells("D1:E1")  # Modernize section
            summary_sheet.merge_cells("F1:G1")  # Modernize & Downsize section

        post_auto_adjustment(summary_sheet)
    except Exception as e:
        log_message(LevelType.INFO, f"[translate:create_summary_sheet error]: {str(e)}", ErrorCode=-1)

def format_numeric_cells(sheet):
    """
    Applies dollar formatting to columns with 'cost', 'savings', or 'price'
    in their headers, accommodating a variable number of header rows.
    """
    try:
        currency_keywords = ("cost", "price", "saving")
        currency_cols = set()

        # Determine the start of the data rows based on the sheet title.
        if sheet.title == "Total Annual Savings":
            # The summary sheet has a 2-row header.
            data_start_row = 3
        elif sheet.title == "Legal Disclaimer":
            # This sheet has no numeric data to format.
            return
        else:
            # All other sheets, like "Recommended-Instance", have a 4-row header.
            data_start_row = 5
        
        header_end_row = data_start_row - 1

        # Find the columns that contain currency keywords in their headers.
        for row_idx in range(1, header_end_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # Check the value, handling cases where it might be None or a merged cell.
                cell_value = str(cell.value).lower() if cell.value is not None else ""
                if any(keyword in cell_value for keyword in currency_keywords):
                    currency_cols.add(col_idx)

        # Apply currency formatting to the data cells in the identified columns.
        for row_idx in range(data_start_row, sheet.max_row + 1):
            for col_idx in currency_cols:
                cell = sheet.cell(row=row_idx, column=col_idx)
                try:
                    # Convert value to a numeric type if possible
                    cell.value = float(cell.value)
                    # Apply currency format
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                except (ValueError, TypeError):
                    # Do nothing if the cell value is not a number
                    pass
    except Exception as e:
        log_message(LevelType.ERROR, f"format_numeric_cells error: {str(e)}", ErrorCode=-1)

def highlight_recommendations(sheet, run_type):
    """
    Highlights the Recommendation I/II/III Instance columns if the corresponding
    Monthly Savings I/II/III columns have non-negative values.
    """
    try:
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        
        # Find relevant columns
        target_columns = {}
        t_hi = 0
        for i in range(len(sheet[3])):
            cell1 = sheet[3][i].value
            cell2 = sheet[4][i].value
            b_header = ""
            if cell1 is not None and cell1 != "":
                t_header = cell1
                t_hi = i
            if cell2 is not None and cell2 != "":
                b_header = cell2
            target_columns[f"{t_header}/{b_header}"] = (t_hi, i)
        if run_type == "EIA":
            header_col = {'OPTIMAL/Monthly Savings': 'OPTIMAL/Instance', 'GOOD/Monthly Savings': 'GOOD/Instance'}
        elif run_type == "CCA":
            header_col = {"Hourly Cost Optimization/Annual Savings ($)": "Hourly Cost Optimization/Instance", "Modernize/Annual Savings ($)": "Modernize/Instance", "Modernize & Downsize/Annual Savings ($)": "Modernize & Downsize/Instance"}
        else:
            return None

        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            if any(cell.value and isinstance(cell.value, str) and "Grand Total" in cell.value for cell in row):
                continue  # Skip rows containing 'Grand Total'
            for key, val in header_col.items():
                savings_col = target_columns.get(key)[1]
                recommend_col = target_columns.get(val)[1]
                
                if savings_col and recommend_col:
                    savings_cell = row[savings_col]
                    recommend_cell = row[recommend_col]
                    try:
                        savings_value = float(savings_cell.value)
                        if savings_value >= 0:
                            recommend_cell.fill = green_fill
                    except (ValueError, TypeError):
                        pass
    except Exception as e:
        log_message(LevelType.ERROR, f"[translate:highlight_recommendations error]: {str(e)}", ErrorCode=-1)


def post_auto_adjustment(sheet):
    """
    Adjusts column widths, applies formatting, and handles merged headers.
    """
    try:
        max_lengths = {}

        # Detect merged header columns
        merged_header_cells = set()
        for merged_cell in sheet.merged_cells.ranges:
            
            if merged_cell.min_row in (1,2):  # Only handling merged headers in the first row
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    merged_header_cells.add(col)

        border = Border(left=Side(style='thin', color='D3D3D3'),
                        right=Side(style='thin', color='D3D3D3'),
                        top=Side(style='thin', color='D3D3D3'),
                        bottom=Side(style='thin', color='D3D3D3'))

        # Process all columns
        for column in sheet.columns:
            max_length = 10
            column = list(column)  # Convert to list for easier iteration
            for cell in column:

                if cell.value == "Unnamed: 0":
                    cell.value = None  # Set to None (blank in Excel)

                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Apply formatting based on header or data
                if cell.row in (1, 2):
                    # Headers: Apply black fill and white bold text
                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    if len(sheet[1])>10: # indicates not the total annual saving sheet
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif cell.row in (3, 4) and cell.column in merged_header_cells and len(sheet[1])>10:
                    # Headers: Apply black fill and white bold text
                    cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    # Data: Apply white fill and black text
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    cell.font = Font(color="000000")

                # Calculate column width based on content
                if cell.value:
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))

            # Store max length for width adjustment
            if column[0].coordinate:  # Use coordinate instead of column_letter for merged cells
                if isinstance(column[0], MergedCell):
                    col_letter = column[0].coordinate.split('1')[0]  # Extract column letter from coordinate
                else:
                    col_letter = column[0].column_letter

                max_lengths[col_letter] = max_length


        # Adjust column widths
        for col_letter, max_length in max_lengths.items():
            sheet.column_dimensions[col_letter].width = max_length + 2

        format_numeric_cells(sheet)
    except Exception as e:
        log_message(LevelType.ERROR, f"post_auto_adjustment error: {str(e)}", ErrorCode=-1)

def convert(csv_file_list, sheet_names, excel_file, run_type, results_path):
    """
    Converts a list of CSV files into an Excel file with separate sheets named after the CSV files.

    Args:
    - csv_file_list (list): List of paths to CSV files.
    - sheet_names (list): List of sheet names corresponding to each CSV file
    - excel_file (str): Path to the output Excel file.
    """
    try:
        # Function to read CSV and return a list of lists
        def read_csv_to_list(csv_file):
            with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
                reader = csv.reader(file)
                data = list(reader)
            return data
        # Create a new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet created by Workbook
        for i in range(len(csv_file_list)):
            csv_file = csv_file_list[i]
            sheet_name = sheet_names[i]

            # Read data from CSV file
            data = read_csv_to_list(csv_file)

            # For EIA, detect the index of 'Input Headroom' so we can drop that column for all rows
            input_headroom_col_idx = None
            input_headroom_value = None
            if run_type == "EIA" and data and len(data[0]) > 0:
                for idx, hdr in enumerate(data[0]):
                    if isinstance(hdr, str) and hdr.strip() == "Input Headroom":
                        input_headroom_col_idx = idx
                        input_headroom_value = round(float(data[1][idx]), 2)
                        break

            # Create a new sheet and add data from the CSV file
            ws = wb.create_sheet(title=sheet_name)
            
            for row_index, row in enumerate(data):
                cleaned_row = [cell.strip() if isinstance(cell, str) else cell for cell in row]

                # If EIA and Input Headroom column exists, remove it from this row
                if run_type == "EIA" and input_headroom_col_idx is not None and input_headroom_col_idx < len(cleaned_row):
                    cleaned_row.pop(input_headroom_col_idx)

                while cleaned_row and (cleaned_row[-1] == '' or cleaned_row[-1] is None):
                    cleaned_row.pop()  # Remove trailing empty columns
                    
                perf_headers = ["Perf Enhancement I", "Perf Enhancement II", "Perf Enhancement III"]
                perf_indices = [i for i, h in enumerate(data[0]) if h.strip() in perf_headers]
                for idx in perf_indices:
                    try:
                        if row_index > 0:
                            cleaned_row[idx] = round(float(cleaned_row[idx]), 2)
                    except Exception:
                        pass

                # Append cleaned data to the worksheet
                ws.append(cleaned_row)
                if row_index == 0 and run_type == "EIA":
                    for col_index, header in enumerate(cleaned_row):
                        # Strip spaces from header for comparison
                        stripped_header = header.strip()
                        if stripped_header == "Current Instance Emission":
                            ws.cell(row=1, column=col_index + 1, value="Current Instance Emission (kgco2eq)")
                        elif stripped_header == "Instance Emission I":
                            ws.cell(row=1, column=col_index + 1, value="Instance Emission I (kgco2eq)")
                        elif stripped_header == "Instance Emission II":
                            ws.cell(row=1, column=col_index + 1, value="Instance Emission II (kgco2eq)")
                        elif stripped_header == "Instance Emission III":
                            ws.cell(row=1, column=col_index + 1, value="Instance Emission III (kgco2eq)")
                        elif stripped_header == "STATUS":
                            ws.cell(row=1, column=col_index + 1, value="Remark")
                        elif stripped_header == "Recommendation I Instance":
                            ws.cell(row=1, column=col_index + 1, value="OPTIMAL")
                        elif stripped_header == "Recommendation II Instance":
                            ws.cell(row=1, column=col_index + 1, value="GOOD")

            # Get the header row to identify target columns
            header = data[0]
            target_columns = {
                "Number of Instances": None,
                "Current Monthly Cost": None,
                "Annual Cost": None,
                "Monthly Cost I": None,
                "Annual Cost I (perf scaled)": None,
                "Annual Savings I": None,
                "Perf Enhancement I": None,
                "Monthly Cost II": None,
                "Annual Cost II (perf scaled)": None,
                "Annual Savings II": None,
                "Perf Enhancement II": None,
                "Monthly Cost III": None,
                "Annual Cost III (perf scaled)": None,
                "Annual Savings III": None,
                "Perf Enhancement III": None,
                "Current Monthly Price": None,
                "Current Instance Energy Consumption (kwh)": None,
                "Current Instance Emission": None,
                "Monthly Price I": None,
                "Instance Energy Consumption I (kwh)": None,
                "Instance Emission I": None,
                "Perf Enhancement I": None,
                "Monthly Price II": None,
                "Instance Energy Consumption II (kwh)": None,
                "Instance Emission II": None,
                "Perf Enhancement II": None,
                "Monthly Price III": None,
                "Instance Energy Consumption III (kwh)": None,
                "Instance Emission III": None,
                "Perf Enhancement III": None,
                "Monthly Savings I": None,
                "Monthly Savings II": None,
                "Monthly Savings III": None
            }

            # Find the column indexes for the target columns, stripping spaces from headers
            for col_index, col_name in enumerate(header):
                cleaned_col_name = col_name.strip()  # Strip spaces before comparing
                if cleaned_col_name in target_columns:
                    target_columns[cleaned_col_name] = col_index + 1  # Excel uses 1-based index

            # Add "Grand Total" row at the bottom
            last_row = ws.max_row + 1
            uuid_column_index = 1  # Assuming UUID is in the first column (Column A)

            # Write "Grand Total" in the UUID column at the end
            ws.cell(row=last_row, column=uuid_column_index, value="Grand Total")

            # Calculate the sum for each target column
            for col_name, col_index in target_columns.items():
                if col_index is not None:
                    # Check if the column is one of the "Perf Enhancement" columns
                    if "Perf Enhancement" in col_name:
                        # Calculate average for "Perf Enhancement" columns
                        start_row = 2  # Assuming the first row is the header
                        sum_value = 0
                        valid_count = 0  # Count of valid numeric entries
                        for row in range(start_row, last_row):
                            cell_value = ws.cell(row=row, column=col_index).value
                            try:
                                # Convert the cell value to float if it's numeric
                                cell_value = float(cell_value)
                                valid_count += 1  # Increment count for valid entries
                            except Exception as e:
                                # If it's not numeric, skip it
                                cell_value = 0

                            sum_value += cell_value

                        # Calculate average, avoiding division by zero
                        if valid_count > 0:
                            average_value = sum_value / valid_count
                        else:
                            average_value = 0

                        # Write the computed average in the respective column
                        ws.cell(row=last_row, column=col_index, value=round(average_value, 2))

                    else:
                        # Calculate sum for other columns
                        start_row = 2  # Assuming the first row is the header
                        sum_value = 0
                        for row in range(start_row, last_row):
                            cell_value = ws.cell(row=row, column=col_index).value

                            try:
                                # Convert the cell value to float if it's numeric
                                cell_value = float(cell_value)
                            except Exception as e:
                                # If it's not numeric, skip it
                                cell_value = 0

                            sum_value += cell_value

                        # Write the computed sum in the respective column
                        ws.cell(row=last_row, column=col_index, value=sum_value)
            

            # Find relevant columns
            header = data[0]
            target_columns = {col: idx + 1 for idx, col in enumerate(header)}

            create_summary_sheet(wb, ws, run_type)
            # replace headers only after summary sheet is created
            replace_headers(ws, run_type, input_headroom_value)
            if run_type == "CCA" and ws.title == "Recommended-Instance":
                replace_negative_annual_savings(ws, run_type)
            post_auto_adjustment(ws)
            highlight_recommendations(ws, run_type)
            last_row = ws.max_row + 1
            legend_cell = ws.cell(row=last_row, column=uuid_column_index, value="Note : Green color instances indicate positive savings.")
            red_font = Font(color="0000FF")
            legend_cell.font = red_font

        # Adding the Legal Disclaimer sheet with separate content for CCA and EIA
        if run_type == "EIA":
            disclaimer_content = (
                "Disclaimer: THE MATERIALS PROVIDED THROUGH THIS TOOL ARE PROVIDED 'AS IF', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.\n\n"
                "EPYC Cloud Instance Advisory recommendations generated using AMD EPYC Cloud Instance Advisor\n\n"
                "Copyright - 2025 Advanced Microdevices Inc.\n\n"
                "For Terms of Use / Copyrights: please refer https://www.amd.com/en/legal/copyright.html\n"
            )
        elif run_type == "CCA":
            disclaimer_content = (
                "Disclaimer: THE MATERIALS PROVIDED THROUGH THIS TOOL ARE PROVIDED 'AS IF', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.\n\n"
                "EPYC Cloud Cost Advisory recommendations generated using AMD EPYC Cloud Cost Advisor\n\n"
                "Copyright - 2025 Advanced Microdevices Inc.\n\n"
                "For Terms of Use / Copyrights: please refer https://www.amd.com/en/legal/copyright.html\n"
            )
        else:
            disclaimer_content = (
                "Disclaimer: THE MATERIALS PROVIDED THROUGH THIS TOOL ARE PROVIDED 'AS IF', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.\n\n"
                "Cloud instance recommendation generated using AMD EPYC Cloud Instance Advisor\n\n"
                "Copyright - 2025 Advanced Microdevices Inc.\n\n"
                "For Terms of Use / Copyrights: please refer https://www.amd.com/en/legal/copyright.html\n"
            )

        legal_disclaimer = wb.create_sheet(title="Legal Disclaimer")
        for line in disclaimer_content.split('\n'):
            legal_disclaimer.append([line])
        # Save the workbook
        wb.save(excel_file)
        return excel_file
    except Exception as e:
        log_message(LevelType.ERROR, f"[translate:convert error]: {str(e)}", ErrorCode=-1)
        return None

def replace_headers(sheet, run_type, input_headroom_value):
    """
    Replaces the sheet headers with formatted headers based on the run type.
    - Inserts a title across the first two rows and merges all columns.
    - Aligns the title text to the left.
    """
    try:
        # Define the title row
        title_row_text = f"{run_type.upper()} Recommendations"
        max_col = sheet.max_column  # Get total columns for merging

        # Define Headers Based on Run Type
        if run_type == "EIA":
            title_row_text = f"EPYC Cloud Instance Advisory Recommendations - ( Operational Safety Margin : {input_headroom_value})%"
            first_header_row = ["UUID", "CSP", "Pricing Model", "Zone", "Current Instance", "vCPU", "Current Monthly Price", "Current Instance Energy Consumption (kwh)", "Current Instance Emission (kgco2eq)", "OPTIMAL", "", "", "", "", "", "", "GOOD", "", "", "", "", "", "", "Remark"]
            
            second_header_row = ["UUID", "CSP", "Pricing Model", "Zone", "Current Instance", "vCPU", "Current Monthly Price", "Current Instance Energy Consumption (kwh)", "Current Instance Emission (kgco2eq)", "Instance", "vCPU", "Monthly Price", "Monthly Savings", "Instance Energy Consumption (kwh)", "Instance Emission (kgco2eq)", "Perf Enhancement", "Instance", "vCPU", "Monthly Price", "Monthly Savings", "Instance Energy Consumption (kwh)", "Instance Emission (kgco2eq)", "Perf Enhancement", "Remark"]

            merge_cell_list = ("A3:A4", "B3:B4", "C3:C4", "D3:D4", "E3:E4", "F3:F4", "G3:G4", "H3:H4", "I3:I4", "J3:P3", "Q3:W3", "X3:X4")
        
        elif run_type == "CCA":
            title_row_text = f"EPYC Cloud Cost Advisory Recommendations"
            first_header_row = ['UUID/Instance Name', 'Cloud', 'Pricing Model', 'Region', 'Current Instance', 'Quantity', 'Current vCPU(s)', 'Current Monthly Cost', 'Current Annual Cost', 'Hourly Cost Optimization', '', '', '', '', '', 'Modernize', '', '', '', '', '', 'Modernize & Downsize', '', '', '', '', '', 'Remark']
            
            second_header_row = ['UUID/Instance Name', 'Cloud', 'Pricing Model', 'Region', 'Current Instance', 'Quantity', 'Current vCPU(s)', 'Current Monthly Cost', 'Current Annual Cost', 'Instance', 'vCPU(s)', 'Monthly Cost ($)', 'Annual Cost ($)', 'Annual Savings ($)', 'Performance Improvement', 'Instance', 'vCPU(s)', 'Monthly Cost ($)', 'Annual Cost ($)', 'Annual Savings ($)', 'Performance Improvement', 'Instance', 'vCPU(s)', 'Monthly Cost ($)', 'Annual Cost ($)', 'Annual Savings ($)', 'Performance Improvement', 'Remark']

            merge_cell_list = ("A3:A4", "B3:B4", "C3:C4", "D3:D4", "E3:E4", "F3:F4", "G3:G4", "H3:H4", "I3:I4", "J3:O3", "P3:U3", "V3:AA3", "AB3:AB4")

        # Delete Original Header and Insert 4 New Rows
        sheet.delete_rows(1)  
        sheet.insert_rows(1, 4)  

        # Insert Title Rows & Merge Cells
        title_cell = sheet.cell(row=1, column=1)
        title_cell.value = title_row_text
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")  # Left align
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=max_col)

        # Adjust alignment for all merged title cells
        for col in range(1, max_col + 1):
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal="left", vertical="center")
            sheet.cell(row=2, column=col).alignment = Alignment(horizontal="left", vertical="center")
        # Write First Header Row
        for col_index, value in enumerate(first_header_row, start=1):
            sheet.cell(row=3, column=col_index).value = value

        # Write Second Header Row
        for col_index, value in enumerate(second_header_row, start=1):
            sheet.cell(row=4, column=col_index).value = value

        # Merge Cells for Multi-Row Headers
        for merge_cell_range in merge_cell_list:
            sheet.merge_cells(merge_cell_range)
    except Exception as e:
        log_message(LevelType.ERROR, f"[translate:replace_headers error]: {str(e)}", ErrorCode=-1)
        
        
def _build_grouped_header_map(sheet):
    """
    Build map: "Group/Leaf" -> column_index (0-based for row tuples).
    Uses row 3 for group headers and row 4 for leaf headers.
    """
    mapping = {}
    current_group = ""
    row3 = sheet[3]
    row4 = sheet[4]

    for idx in range(len(row3)):
        top = row3[idx].value
        leaf = row4[idx].value

        if top is not None and top != "":
            current_group = str(top)

        leaf_text = str(leaf) if leaf is not None and leaf != "" else ""
        mapping[f"{current_group}/{leaf_text}"] = idx

    return mapping


def _resolve_savings_columns(header_map):
    """
    Return list of 0-based indices for the three Annual Savings columns.
    """
    keys = [
        "Hourly Cost Optimization/Annual Savings ($)",
        "Modernize/Annual Savings ($)",
        "Modernize & Downsize/Annual Savings ($)",
    ]
    cols = [header_map[k] for k in keys if k in header_map]
    return cols


def _is_grand_total_row(row):
    """
    True if any cell contains 'Grand Total' as a substring.
    """
    for c in row:
        v = c.value
        if isinstance(v, str) and "Grand Total" in v:
            return True
    return False


def _to_float(v):
    """
    Safe float conversion; returns None if not convertible.
    Avoids using exceptions for core control flow paths.
    """
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _replace_negatives_in_columns(sheet, col_indices):
    """
    Walk data rows and replace negative numeric values with text, set text format.
    """
    for row in sheet.iter_rows(min_row=5, max_row=sheet.max_row):
        if _is_grand_total_row(row):
            continue
        for col_idx in col_indices:
            cell = row[col_idx]
            val = _to_float(cell.value)
            if val is not None and val < 0:
                cell.value = NEGATIVE_REPLACEMENT_TEXT
                cell.number_format = numbers.FORMAT_TEXT


def replace_negative_annual_savings(sheet, run_type):
    """
    Replace negative Annual Savings values with a fixed text in CCA sheets.
    Assumes a two-row header at rows 3-4 and data from row 5 onward.
    """
    if run_type != "CCA":
        return  # Only applicable to CCA

    header_map = _build_grouped_header_map(sheet)
    savings_cols = _resolve_savings_columns(header_map)

    if not savings_cols:
        return

    _replace_negatives_in_columns(sheet, savings_cols)


def csv_to_excel_generation(csv_files, sheet_names, output_excel_file, run_type, results_path):
    try:
        sheet_names = sheet_names.split(",")

        # Truncate sheet names to 31 characters and remove invalied charecters
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for i in range(len(sheet_names)):
            sheet_name = sheet_names[i]
            flag=False
            given_sheet_name = sheet_name
            for char in invalid_chars:
                if char in sheet_name:
                    flag=True
                    sheet_name = sheet_name.replace(char, '')
            if flag:
                print(f"invalid characters in the sheet name: {given_sheet_name}")
            sheet_names[i] = sheet_name[:31]

        return convert(csv_files, sheet_names, output_excel_file, run_type, results_path)
    except Exception as e:
        log_message(LevelType.ERROR, f"Excel generation failed: {str(e)}", ErrorCode=-1)
