from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.connections.pylogger import log_message
from app.utils.constants import LevelType


BLACK_HEADER_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

WHITE_BOLD_FONT = Font(bold=True, color="FFFFFF")
NORMAL_FONT = Font(bold=False, color="000000")
NOTE_FONT = Font(name='Cambria', size=11, color="0000FF", italic=False)
BOLD_NORMAL_FONT = Font(bold=True, color="000000")
DISCLAIMER_FONT = Font(name='Cambria', size=11, bold=False, color="000000")
DISCLAIMER_TITLE_FONT = Font(name='Cambria', size=11, bold=False, color="000000")

HEADER_ALIGNMENT = Alignment(wrap_text=True, horizontal="center", vertical="center")
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")

WHITE_SIDE = Side(border_style="thin", color="FFFFFF")
BLACK_SIDE = Side(border_style="thin", color="000000")
THIN_BLACK_SIDE = Side(border_style="thin", color="000000")

# Defined Borders
R1_BOTTOM_BORDER = Border(bottom=WHITE_SIDE, left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE)
R4_HEADER_BORDER = Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=None)
R5_BOTTOM_BORDER = Border(left=WHITE_SIDE, right=WHITE_SIDE, top=None, bottom=WHITE_SIDE) 
TITLE_BORDER = Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE)


THICK_BLACK_RIGHT_SIDE = Side(border_style="medium", color="000000")
R3_SEPARATOR_BORDER = Border()
R4_SEPARATOR_BORDER = Border()

# --- New/Modified Global Constants for Column Widths ---
INSTANCE_TYPE_COLS = [1, 10, 17] 
SMALL_INSTANCE_TYPE_WIDTH = 12 
MAX_COLUMN_WIDTH = 35 # Set the overall maximum width limit for all columns



def apply_cell_style(cell, fill=None, font=None, alignment=None, border=None):
    """Applies multiple styles to an openpyxl cell."""
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border


def safe_float(value):
    """Safely converts a value to float, defaulting to 0.0 on error."""
    try: return float(value)
    except (ValueError, TypeError): return 0.0


def safe_round(value, decimals=2):
    try:
        if value is None or str(value).strip() == "":
            return value
        val = float(value)
        return round(val, decimals)
    except (ValueError, TypeError):
        return value


def create_disclaimer_sheet(wb):
    ws = wb.create_sheet("Legal Disclaimer")
    
    ws.column_dimensions['A'].width = 150
    
    ws.cell(1, 1, value="Disclaimer: THE MATERIALS PROVIDED THROUGH THIS TOOL ARE PROVIDED 'AS IF', WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.")
    apply_cell_style(ws.cell(1, 1), font=DISCLAIMER_TITLE_FONT, alignment=LEFT_ALIGNMENT)
    
    ws.cell(3, 1, value="EPYC Cloud Instance Advisory recommendations generated using AMD EPYC Cloud Instance Advisor")
    apply_cell_style(ws.cell(3, 1), font=DISCLAIMER_FONT, alignment=LEFT_ALIGNMENT)

    current_year = datetime.now().year
    ws.cell(5, 1, value=f"Copyright - {current_year} Advanced Microdevices Inc.")
    apply_cell_style(ws.cell(5, 1), font=DISCLAIMER_FONT, alignment=LEFT_ALIGNMENT)
    
    ws.cell(7, 1, value="For Terms of Use / Copyrights: please refer https://www.amd.com/en/legal/copyright.html")
    apply_cell_style(ws.cell(7, 1), font=DISCLAIMER_FONT, alignment=LEFT_ALIGNMENT)
    
    return ws


def create_total_savings_sheet(wb, grand_total):
    ws = wb.create_sheet("Total Annual Savings")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20
    
    COL_CURRENT_COST = 1
    COL_OPT_COST = 2
    COL_OPT_SAVINGS = 3
    COL_GOOD_COST = 4
    COL_GOOD_SAVINGS = 5
    TOTAL_COLS = 5

    
    ws.merge_cells(start_row=1, start_column=COL_CURRENT_COST, end_row=2, end_column=COL_CURRENT_COST)
    current_cost_cell = ws.cell(1, COL_CURRENT_COST, value="Current Cost")
    apply_cell_style(current_cost_cell, BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE))
    
    ws.merge_cells(start_row=1, start_column=COL_OPT_COST, end_row=1, end_column=COL_OPT_SAVINGS)
    optimal_cell = ws.cell(1, COL_OPT_COST, value="OPTIMAL")
    apply_cell_style(optimal_cell, BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE))
    
    ws.merge_cells(start_row=1, start_column=COL_GOOD_COST, end_row=1, end_column=COL_GOOD_SAVINGS)
    good_cell = ws.cell(1, COL_GOOD_COST, value="GOOD")
    apply_cell_style(good_cell, BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE))

    header_row_2 = {
        COL_OPT_COST: "Total Cost", COL_OPT_SAVINGS: "Total Savings",
        COL_GOOD_COST: "Total Cost", COL_GOOD_SAVINGS: "Total Savings"
    }
    for col_idx, header_text in header_row_2.items():
        cell = ws.cell(2, col_idx, value=header_text)
        apply_cell_style(cell, BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, Border(left=WHITE_SIDE, right=WHITE_SIDE, top=WHITE_SIDE, bottom=WHITE_SIDE))

    # Multiply by 12 to convert monthly to annual
    data_map = {
        COL_CURRENT_COST: safe_round(grand_total.get("Current Monthly Price", 0.0) * 12),
        COL_OPT_COST: safe_round(grand_total.get("Monthly Price I", 0.0) * 12),
        COL_OPT_SAVINGS: safe_round(grand_total.get("Monthly Savings I", 0.0) * 12), 
        COL_GOOD_COST: safe_round(grand_total.get("Monthly Price II", 0.0) * 12),
        COL_GOOD_SAVINGS: safe_round(grand_total.get("Monthly Savings II", 0.0) * 12),
    }

    data_row = 3
    for col_idx in range(1, TOTAL_COLS + 1):
        cell = ws.cell(data_row, col_idx, value=data_map.get(col_idx))
        
        # Changed from GRAY_FILL to WHITE_FILL and alignment from RIGHT to CENTER
        cell.fill = WHITE_FILL
        cell.number_format = '"$"#,##0.00'
        apply_cell_style(cell, alignment=CENTER_ALIGNMENT, font=BOLD_NORMAL_FONT)
        
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    return ws


def create_main_sheet(wb, data, grand_total, safety_margin):
    """Creates and populates the main sheet with the requested 5-row structure."""
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
        
    ws = wb.create_sheet(title="Recommended-Instance", index=0)

    TOTAL_COLUMNS = 26
    

    ROW_TITLE_START = 1 
    ROW_TITLE_END = 2 
    ROW_GROUP_HEADER = 3 
    ROW_DETAIL_HEADER = 4 
    ROW_DATA_START = 5 

    ws.row_dimensions[ROW_TITLE_START].height = 15
    ws.row_dimensions[ROW_TITLE_END].height = 10 
    ws.row_dimensions[ROW_GROUP_HEADER].height = 15
    ws.row_dimensions[ROW_DETAIL_HEADER].height = 35
    
    

    ws.merge_cells(start_row=ROW_TITLE_START, start_column=1, end_row=ROW_TITLE_END, end_column=TOTAL_COLUMNS) 
    cell_a1 = ws["A1"]
    cell_a1.value = f"EPYC Cloud Instance Advisory Recommendations - ( Operational Safety Margin : {safety_margin} )%"
    apply_cell_style(
        cell_a1, 
        BLACK_HEADER_FILL, 
        WHITE_BOLD_FONT, 
        Alignment(horizontal="left", vertical="center"), 
        TITLE_BORDER
    )


    for row_idx in [ROW_GROUP_HEADER, ROW_DETAIL_HEADER]:
        for col_idx in range(1, TOTAL_COLUMNS + 1):
            ws.cell(row_idx, col_idx).fill = BLACK_HEADER_FILL
            ws.cell(row_idx, col_idx).border = Border()
    
    
    ws.merge_cells(start_row=ROW_GROUP_HEADER, start_column=1, end_row=ROW_GROUP_HEADER, end_column=8)
    apply_cell_style(ws[f"A{ROW_GROUP_HEADER}"], BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, R5_BOTTOM_BORDER) 


    ws.merge_cells(start_row=ROW_GROUP_HEADER, start_column=11, end_row=ROW_GROUP_HEADER, end_column=18)
    cell_k3 = ws[f"K{ROW_GROUP_HEADER}"]
    cell_k3.value = "Optimal" 


    apply_cell_style(cell_k3, BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, R1_BOTTOM_BORDER) 


    ws.merge_cells(start_row=ROW_GROUP_HEADER, start_column=19, end_row=ROW_GROUP_HEADER, end_column=26)
    ws[f"S{ROW_GROUP_HEADER}"].value = "Good" 
    apply_cell_style(ws[f"S{ROW_GROUP_HEADER}"], BLACK_HEADER_FILL, WHITE_BOLD_FONT, CENTER_ALIGNMENT, R1_BOTTOM_BORDER) 
    
    

    header_map_final = [
        # Current (1-8)
        (1, "Region", "region", "current"),
        (2, "Instance Type", "type", "current"), (3, "Cost ($)", "cost", "current"), (4, "Power (kW)", "power", "current"), 
        (5, "Carbon (kgCO2eq)", "carbon", "current"), (6, "UUID/Instance Name", "id", "top"), (7, "Cloud", "csp", "top"), 
        (8, "Pricing Model", "pricingModel", "current"), 
        
        (9, "vCPU(s)", "vCPU", "current_vcpus_for_remark"), 
        (10, "Remark", "status", "current"), 
        (11, "Instance Type", "type", 0), # COL 11 is Instance Type (Optimal)
        (12, "vCPU(s)", "vCPU", 0), (13, "Cost ($)", "cost", 0), 
        (14, "Power (kW)", "power", 0), (15, "Carbon (kgCO2eq)", "carbon", 0), (16, "Monthly Savings ($)", "monthlySavings", 0), 
        (17, "Performance Improvement", "perf", 0), (18, "Untapped Capacity", "untappedCapacity", 0),
        
        (19, "Instance Type", "type", 1), # COL 19 is Instance Type (Good)
        (20, "vCPU(s)", "vCPU", 1), (21, "Cost ($)", "cost", 1), 
        (22, "Power (kW)", "power", 1), (23, "Carbon (kgCO2eq)", "carbon", 1), (24, "Monthly Savings ($)", "monthlySavings", 1), 
        (25, "Performance Improvement", "perf", 1), (26, "Untapped Capacity", "untappedCapacity", 1),
    ]

    for col_idx, header_text, *rest in header_map_final:
        cell = ws.cell(ROW_DETAIL_HEADER, col_idx, value=header_text)
        

        apply_cell_style(cell, BLACK_HEADER_FILL, WHITE_BOLD_FONT, HEADER_ALIGNMENT, R4_HEADER_BORDER)
            
 

    
    row_num = ROW_DATA_START
    for item in data:
        current = item["data"]["currentPlatform"]
        recs = item["data"]["recommendations"]
        row_data = [""] * TOTAL_COLUMNS

        rec1_savings = safe_float(recs[0]["monthlySavings"]) if len(recs) > 0 and recs[0].get("monthlySavings") else 0.0
        rec2_savings = safe_float(recs[1]["monthlySavings"]) if len(recs) > 1 and recs[1].get("monthlySavings") else 0.0

        # Populate Data Row
        for col_idx, _, key, source_type in header_map_final:
            value = "-"
            
            if source_type == "top":
                value = str(item.get(key, "-"))
            elif source_type == "current":
                val = current.get(key, "-")
                if key in ["cost", "power", "carbon", "vCPU"]:
                    value = safe_round(val)
                else:
                    value = str(val)
            elif source_type == "current_vcpus_for_remark":
                value = safe_round(current.get("vCPU", "-"))
            elif isinstance(source_type, int) and len(recs) > source_type:
                val = recs[source_type].get(key, "-")
                if key in ["cost", "power", "carbon", "vCPU", "monthlySavings", "perf", "untappedCapacity"]:
                    value = safe_round(val)
                else: 
                    value = str(val)
            
            row_data[col_idx - 1] = value

        ws.append(row_data)

        data_row_idx = row_num 
        for col_idx in range(1, TOTAL_COLUMNS + 1):
            cell = ws.cell(data_row_idx, col_idx)
            
            is_green = False
            # Check for Optimal Instance Type (Col 11)
            if col_idx == 11 and rec1_savings > 0:
                is_green = True
            # Check for Good Instance Type (Col 19)
            elif col_idx == 19 and rec2_savings > 0:
                is_green = True
            
            if is_green:
                cell.fill = LIGHT_GREEN_FILL
                
            if col_idx in [3, 4, 5, 13, 14, 15, 16, 21, 22, 23, 24]:
                cell.number_format = '#,##0.00'

            cell.alignment = LEFT_ALIGNMENT
            cell.border = Border()
                
        row_num += 1

    gt_row = [""] * TOTAL_COLUMNS 
    
    gt_row[0] = "Grand Total"

    gt_map = {
        3: "Current Monthly Price", 4: "Current Instance Energy Consumption (kwh)", 5: "Current Instance Emission",
        13: "Monthly Price I", 14: "Instance Energy Consumption I (kwh)", 15: "Instance Emission I", 16: "Monthly Savings I", 17: "Perf Enhancement I", 18: "Untapped Capacity I",
        21: "Monthly Price II", 22: "Instance Energy Consumption II (kwh)", 23: "Instance Emission II", 24: "Monthly Savings II", 25: "Perf Enhancement II", 26: "Untapped Capacity II",
    }
    
    for col_idx, key in gt_map.items():
        if key in grand_total:
            gt_row[col_idx - 1] = safe_round(grand_total[key])

    ws.append(gt_row) 
    last_row = ws.max_row 
    
    for col_idx in range(1, TOTAL_COLUMNS + 1):
        cell = ws.cell(last_row, col_idx)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT_ALIGNMENT

        if col_idx in gt_map.keys():
            try:
                # Apply number formatting 
                if col_idx in [3, 13, 16, 21, 24]:  # Cost and Savings columns
                    cell.number_format = '"$"#,##0.00'
                else:
                    cell.number_format = '#,##0.00'
            except:
                pass

        if col_idx == 18 or col_idx == 26:
            cell.border = Border()


    # --- Note/Legend ---
    note_text = "Note : Green color instances indicate positive savings."
    note_row = ws.max_row + 2
    ws.merge_cells(f'A{note_row}:F{note_row}')
    note_cell = ws[f'A{note_row}']
    note_cell.value = note_text
    apply_cell_style(note_cell, font=NOTE_FONT, alignment=LEFT_ALIGNMENT)
    
    # Add second note
    note_row2 = note_row + 1
    ws.merge_cells(f'A{note_row2}:F{note_row2}')
    note_cell2 = ws[f'A{note_row2}']
    note_cell2.value = "Sizing instances - matching resources to actual demand (if Applicable)"
    apply_cell_style(note_cell2, font=NOTE_FONT, alignment=LEFT_ALIGNMENT)
    
    ws.freeze_panes = "F5"

    return ws


def adjust_column_widths(ws):
    """
    Adjusts column widths for the main sheet based on content, 
    applying a maximum width limit and skipping the 'Instance Type' columns.
    """
    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)

        # Skip Instance Type columns to keep their fixed small width
        if col_idx in [2, 11, 19]: 
            continue 

        max_len = 0
        
        # Check from the header row (Row 4)
        for cell in col:
            if cell.value and cell.row >= 4: 
                # Use a factor of 1.2 to account for font size/padding, but use 1.0 for UUID/Instance Name (Col 6)
                # as it often contains long non-readable strings that can be cut off.
                multiplier = 1.0 if col_idx == 6 else 1.2
                current_len = len(str(cell.value)) * multiplier
                max_len = max(max_len, current_len)

        if max_len > 0:
            # Apply auto-adjustment, respecting the MAX_COLUMN_WIDTH
            adjusted_width = max(16, max_len)
            ws.column_dimensions[col_letter].width = min(MAX_COLUMN_WIDTH, adjusted_width)


def generate_excel_report(data, grand_total, output_path, headroom):
    """Main function to generate the complete Excel report with three tabs."""
    try:
        if not data and not grand_total:
            log_message(LevelType.ERROR, "Aborting Excel generation due to data loading errors.", ErrorCode=-1)
            return

        wb = Workbook()
        
        ws_main = create_main_sheet(wb, data, grand_total, headroom)
        ws_totals = create_total_savings_sheet(wb, grand_total)
        ws_disclaimer = create_disclaimer_sheet(wb)
        
        wb.active = ws_main
        
        # Adjust widths only after all content is written
        adjust_column_widths(ws_main)

        wb.save(output_path)
        log_message(LevelType.INFO, f"Excel successfully generated at: {output_path}")
    except Exception as e:
        log_message(LevelType.ERROR, f"Error generating or saving Excel file: {e}", ErrorCode=-1)
